"""
Processador de Relatórios — Sólides + Bradesco Dental
Interface web para upload de PDFs e download de Excel

Dependências:
    pip install flask openpyxl

Sistema:
    pdftotext (poppler-utils):  sudo apt install poppler-utils

Uso:
    python app.py
    Acesse: http://localhost:5000
"""

import io
import re
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template_string, request, send_file, jsonify

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _extract_text_with_pypdf2(pdf_path: str, password: str | None = None) -> str:
    if PdfReader is None:
        return ""

    reader = PdfReader(pdf_path)
    if reader.is_encrypted:
        if password is None:
            password = ""
        try:
            result = reader.decrypt(password)
        except Exception:
            result = 0
        if result == 0:
            raise RuntimeError("PDF protegido por senha. Digite a senha correta ou envie um PDF sem proteção.")

    text_parts = []
    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        text_parts.append(page_text)
    return "\n".join(text_parts)


def pdf_to_text(pdf_bytes: bytes, password: str | None = None) -> str:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    txt_path = tmp_path.replace(".pdf", ".txt")

    def run_pdftotext(args: list[str]) -> subprocess.CompletedProcess:
        return subprocess.run(args, capture_output=True, text=True)

    try:
        pdftotext_cmd = shutil.which("pdftotext")
        text = ""

        if pdftotext_cmd:
            if password:
                for args in [
                    [pdftotext_cmd, "-layout", "-upw", password, tmp_path, txt_path],
                    [pdftotext_cmd, "-layout", "-opw", password, tmp_path, txt_path],
                ]:
                    r = run_pdftotext(args)
                    if r.returncode == 0:
                        with open(txt_path, encoding="utf-8") as f:
                            text = f.read()
                        break
            if not text:
                args = [pdftotext_cmd, "-layout", tmp_path, txt_path]
                r = run_pdftotext(args)
                if r.returncode == 0:
                    with open(txt_path, encoding="utf-8") as f:
                        text = f.read()

        if not text and PdfReader is not None:
            text = _extract_text_with_pypdf2(tmp_path, password)

        if not text:
            raise RuntimeError(
                "Não foi possível extrair texto do PDF. Verifique se o arquivo está protegido e se a senha 08423 está correta."
            )

        return text
    finally:
        for p in [tmp_path, txt_path]:
            try:
                os.unlink(p)
            except Exception:
                pass


def hms_to_min(s: str) -> int:
    if not s or ":" not in s:
        return 0
    s = s.strip()
    neg = s.startswith("-")
    s = s.lstrip("-")
    parts = s.split(":")
    if len(parts) != 2:
        return 0
    try:
        mins = int(parts[0]) * 60 + int(parts[1])
    except ValueError:
        return 0
    return -mins if neg else mins


def min_to_hms(m: int) -> str:
    neg = m < 0
    m = abs(m)
    h, mn = divmod(m, 60)
    return f"{'-' if neg else ''}{h:02d}:{mn:02d}"


def make_border(color="C0C0C0"):
    t = Side(style="thin", color=color)
    return Border(left=t, right=t, top=t, bottom=t)


# ─────────────────────────────────────────────────────────────────────────────
# PARSER — SÓLIDES
# ─────────────────────────────────────────────────────────────────────────────

THRESHOLD_MIN = 72  # 1:12 per day in minutes


def _extract_window(line: str, start: int, end: int) -> str:
    """Extract first HH:MM value within a character window of a line."""
    segment = line[start:end] if len(line) > start else ""
    m = re.search(r"-?\d{1,3}:\d{2}", segment)
    return m.group() if m else ""


def _get_header_positions(page: str) -> tuple[dict, str]:
    """Return column positions dict and regime ('normal' or 'banco') for a page."""
    for line in page.split("\n"):
        if "TRABALHADAS" in line and "DIA" in line:
            pos = {}
            for col in ["TRABALHADAS", "INTRAJORNADA", "ABONO", "PREVISTAS",
                        "ATRASO", "EXTRAS", "SALDO", "FALTAS"]:
                idx = line.find(col)
                if idx >= 0:
                    pos[col] = idx
            regime = "banco" if "SALDO" in pos and "EXTRAS" not in pos else "normal"
            return pos, regime
    return {}, "normal"


def _classify_day_hours(page: str) -> tuple[int, int]:
    """
    Walk every day line and classify extras per day:
      - day extra > 1:12  → hour_extra bucket
      - 0 < day extra ≤ 1:12 → banco bucket
    Returns (extra_minutes, banco_minutes).
    """
    header_pos, regime = _get_header_positions(page)
    if not header_pos:
        return 0, 0

    extras_col = header_pos.get("EXTRAS", header_pos.get("SALDO", 178))
    extra_total = 0
    banco_total = 0
    atraso_total = 0
    atraso_pos = header_pos.get("ATRASO", 165)

    for line in page.split("\n"):
        if not re.search(r"^\s+\d{2}/\d{2}\s+\w", line):
            continue
        if re.search(r"FERIADO|feriado", line, re.I):
            continue
        if re.match(r"\s+\d{2}/\d{2}\s+(?:sabado|domingo)\s+\-", line):
            continue

        val = _extract_window(line, extras_col - 2, extras_col + 15)
        mins = hms_to_min(val)

        if mins > THRESHOLD_MIN:
            extra_total += mins
        elif mins > 0:
            banco_total += mins

        # Capture atraso (regime normal only — negative value in ATRASO column)
        if regime == "normal":
            atr_val = _extract_window(line, atraso_pos - 2, atraso_pos + 12)
            atr_min = hms_to_min(atr_val)
            if atr_min < 0:
                atraso_total += atr_min  # already negative

    return extra_total, banco_total, atraso_total


def parse_solides(pdf_bytes: bytes, password: str | None = None) -> bytes:
    text = pdf_to_text(pdf_bytes, password)
    pages = text.split("\f")

    records = []
    for page in pages:
        if "DADOS DO COLABORADOR" not in page or "CPF:" not in page:
            continue
        r = _parse_solides_page(page)
        if r:
            records.append(r)

    if not records:
        raise RuntimeError(
            "Nenhum registro Sólides encontrado. Verifique se o PDF é de folha de ponto Sólides e se a senha está correta."
        )

    return _build_solides_excel(records)


def _parse_solides_page(page: str) -> dict | None:
    def get(pattern, text, grp=1, default=""):
        m = re.search(pattern, text)
        return m.group(grp).strip() if m else default

    empresa = get(r"Nome:\s{5,}(.+?)\s{3,}CNPJ:", page)
    nome    = get(r"Nome:\s{2,}(.+?)\s{3,}CPF:", page)
    if not nome:
        return None

    cpf_raw = get(r"CPF:\s+(\d{11})", page)
    cpf = f"{cpf_raw[:3]}.{cpf_raw[3:6]}.{cpf_raw[6:9]}-{cpf_raw[9:]}" if len(cpf_raw) == 11 else cpf_raw

    func_m = re.search(r"Função:\s{2,}(\S.+?)\s{3,}Centro de Custo", page)
    funcao = func_m.group(1).strip() if func_m and "Centro de Custo" not in func_m.group(1) else ""

    cc_m = re.search(r"Centro de Custo:\s+(\d+\s*-\s*[^\n]+)", page)
    centro_custo = cc_m.group(1).strip() if cc_m else ""

    local_m = re.search(r"Local:\s+(.+)", page)
    local = local_m.group(1).strip() if local_m else ""

    # Total line: TRABALHADAS  INTRAJORNADA  ABONO  PREVISTAS  [rest]
    total_m = re.search(r"Total:\s+([\d:]+)\s+(-?[\d:]+)\s+([\d:]+)\s+([\d:]+)", page)
    previstas = trabalhadas = abono = intrajornada_total = ""
    if total_m:
        trabalhadas        = total_m.group(1)
        intrajornada_total = total_m.group(2)
        abono              = total_m.group(3)
        previstas          = total_m.group(4)

    # Per-day classification — returns extras, banco, and atraso (negative)
    extra_min, banco_min, atraso_min = _classify_day_hours(page)
    col_extra = min_to_hms(extra_min) if extra_min > 0 else "00:00"
    col_banco = min_to_hms(banco_min) if banco_min > 0 else "00:00"

    # Saldo = extras + banco + atraso (atraso_min already negative → exact match with PDF)
    saldo_min = extra_min + banco_min + atraso_min
    saldo_hms = min_to_hms(saldo_min)

    # Intrajornada negativa days
    neg_count = sum(
        1 for line in page.split("\n")
        if re.search(r"^\s+\d{2}/\d{2}\s+\w", line)
        and re.search(r"\|\s+\d+:\d+\s+-\d+:\d+", line)
    )

    ajustes = page.count("(m)")

    return {
        "empresa": empresa, "nome": nome, "cpf": cpf,
        "funcao": funcao, "centro_custo": centro_custo, "local": local,
        "previstas": previstas, "trabalhadas": trabalhadas,
        "saldo": saldo_hms,
        "intrajornada_total": intrajornada_total,
        "neg_intra_count": neg_count,
        "col_extra": col_extra,
        "col_banco": col_banco,
        "ajustes_manuais": ajustes,
    }


def _build_solides_excel(records: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folha de Ponto"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    alt_fill  = PatternFill("solid", fgColor="D6E4F0")
    wht_fill  = PatternFill("solid", fgColor="FFFFFF")
    neg_fill  = PatternFill("solid", fgColor="FFE0E0")
    ext_fill  = PatternFill("solid", fgColor="E8F5E9")
    bnk_fill  = PatternFill("solid", fgColor="FFF8E1")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=13, color="1F4E79")
    neg_font  = Font(color="CC0000", bold=True, size=10)
    ext_font  = Font(color="1B5E20", bold=True, size=10)
    bnk_font  = Font(color="E65100", bold=True, size=10)
    border    = make_border()
    center    = Alignment(horizontal="center", vertical="center")
    vcenter   = Alignment(vertical="center")

    ws.merge_cells("A1:L1")
    ws["A1"] = "Relatório de Folha de Ponto — Sólides"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    headers = [
        "Empresa", "Nome do Colaborador", "CPF", "Função", "Centro de Custo",
        "Local", "Horas Previstas", "Horas Trabalhadas", "Saldo do Mês",
        "Horas Extras (> 1:12)", "Banco de Horas (≤ 1:12)",
        "Intrajornada Neg.", "Dias Intra. Negativa", "Ajustes Manuais (m)"
    ]
    ws.row_dimensions[2].height = 36
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, r in enumerate(records):
        row = i + 3
        saldo_neg = r["saldo"].startswith("-")
        extra_pos = r["col_extra"] != "00:00"
        banco_pos = r["col_banco"] not in ("00:00", "") and not r["col_banco"].startswith("-")
        base = alt_fill if i % 2 == 0 else wht_fill

        vals = [
            r["empresa"], r["nome"], r["cpf"], r["funcao"], r["centro_custo"],
            r["local"], r["previstas"], r["trabalhadas"], r["saldo"],
            r["col_extra"], r["col_banco"],
            r["intrajornada_total"], r["neg_intra_count"], r["ajustes_manuais"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border

            if col == 9 and saldo_neg:
                c.fill = neg_fill; c.font = neg_font
            elif col == 10 and extra_pos:
                c.fill = ext_fill; c.font = ext_font
            elif col == 11 and banco_pos:
                c.fill = bnk_fill; c.font = bnk_font
            elif col == 12 and r["intrajornada_total"].startswith("-"):
                c.fill = neg_fill; c.font = neg_font
            else:
                c.fill = base
                c.font = Font(size=10)

            c.alignment = center if col >= 7 else vcenter
        ws.row_dimensions[row].height = 17

    last = len(records) + 3
    ws.merge_cells(f"A{last}:F{last}")
    ws[f"A{last}"] = f"Total: {len(records)} colaboradores"
    ws[f"A{last}"].font = Font(bold=True, color="1F4E79", size=10)

    widths = [28, 34, 16, 24, 24, 26, 14, 16, 12, 18, 18, 14, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PARSER — BRADESCO DENTAL
# ─────────────────────────────────────────────────────────────────────────────

CERTIF_RE = re.compile(
    r"^\s*(\d{7}/\d{2})\s+(\S.+?)\s{2,}"
    r"\d{2}/\d{2}/\d{4}\s+(?:MAS|FEM)\s+\w+\s+"
    r"(CONJ|FILH|OUT)?\s*TNDA\s+\S+\s+"
    r"(?:\w{2}\s+)?\d{2}/\d{4}\s+([\d,.]+[-]?)\s+[\d,.]+\s*$"
)
CONT_RE = re.compile(
    r"^\s*TNDA\s+\S+\s+(?:\w{2}\s+)?\d{2}/\d{4}\s+([\d,.]+[-]?)\s+[\d,.]+\s*$"
)


def _parse_val(s: str) -> float:
    s = s.replace(".", "").replace(",", ".")
    return -float(s[:-1]) if s.endswith("-") else float(s)


def parse_bradesco(pdf_bytes: bytes, password: str | None = None) -> bytes:
    text = pdf_to_text(pdf_bytes, password)
    lines = text.splitlines()

    # Pass 1: titular names
    titulares: dict[str, str] = {}
    for line in lines:
        m = CERTIF_RE.match(line)
        if m:
            num, seq = m.group(1).split("/")
            if seq == "00" and m.group(2).strip():
                titulares[num] = m.group(2).strip()

    # Pass 2: accumulate per certif
    records_map: dict[str, dict] = {}
    order: list[str] = []
    current: str | None = None

    for line in lines:
        m = CERTIF_RE.match(line)
        if m:
            certif = m.group(1)
            nome = m.group(2).strip()
            if not nome:
                continue
            paren_raw = m.group(3)
            num, seq_str = certif.split("/")
            seq = int(seq_str)
            par = "Titular" if seq == 0 else {
                "CONJ": "Cônjuge", "FILH": "Filho(a)", "OUT": "Outro"
            }.get(paren_raw, "Dependente")
            tit = "" if seq == 0 else titulares.get(num, "")

            if certif not in records_map:
                records_map[certif] = {
                    "certif": certif, "nome": nome, "parentesco": par,
                    "titular_nome": tit, "valor": 0.0,
                }
                order.append(certif)

            records_map[certif]["valor"] += _parse_val(m.group(4))
            current = certif
        else:
            mc = CONT_RE.match(line)
            if mc and current:
                records_map[current]["valor"] += _parse_val(mc.group(1))

    records = [records_map[c] for c in order]
    if not records:
        raise RuntimeError(
            "Nenhum registro Bradesco Dental encontrado. Verifique se o PDF é uma fatura Bradesco Dental válida e se a senha está correta."
        )
    return _build_bradesco_excel(records)


def _build_bradesco_excel(records: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fatura Dental"

    hdr_fill  = PatternFill("solid", fgColor="1A3C5E")
    tit_fill  = PatternFill("solid", fgColor="D6EAF8")
    dep_fill  = PatternFill("solid", fgColor="FFFFFF")
    alt_fill  = PatternFill("solid", fgColor="EBF5FB")
    neg_fill  = PatternFill("solid", fgColor="FDECEA")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    tit_font  = Font(bold=True, size=10, color="1A3C5E")
    dep_font  = Font(size=10, color="333333")
    neg_font  = Font(bold=True, size=10, color="CC0000")
    border    = make_border()
    center    = Alignment(horizontal="center", vertical="center")
    indent1   = Alignment(vertical="center", indent=1)

    ws.merge_cells("A1:E1")
    ws["A1"] = "Fatura Técnica — Bradesco Dental"
    ws["A1"].font = Font(bold=True, size=13, color="1A3C5E")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ["Certificado", "Nome do Segurado", "Parentesco", "Nome do Titular", "Valor Líquido (R$)"]
    ws.row_dimensions[2].height = 32
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, r in enumerate(records):
        row = i + 3
        seq = int(r["certif"].split("/")[1])
        is_tit = seq == 0
        is_neg = r["valor"] < 0
        fill = neg_fill if is_neg else (tit_fill if is_tit else (alt_fill if i % 2 else dep_fill))

        vals = [r["certif"], r["nome"], r["parentesco"], r["titular_nome"], r["valor"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.border = border
            if col == 5:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = neg_font if is_neg else (tit_font if is_tit else dep_font)
            elif col == 2:
                c.alignment = indent1
                c.font = tit_font if is_tit else dep_font
            else:
                c.alignment = center
                c.font = tit_font if is_tit else dep_font
        ws.row_dimensions[row].height = 17

    last = len(records) + 4
    total_pos = sum(r["valor"] for r in records if r["valor"] > 0)
    total_neg = sum(r["valor"] for r in records if r["valor"] < 0)
    for label, val in [("Total a cobrar:", total_pos), ("Total a devolver:", total_neg), ("Total líquido:", total_pos + total_neg)]:
        ws.merge_cells(f"A{last}:D{last}")
        ws[f"A{last}"] = label
        ws[f"A{last}"].font = Font(bold=True, size=10, color="1A3C5E")
        ws[f"A{last}"].alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=last, column=5, value=val)
        c.number_format = "#,##0.00"
        c.font = Font(bold=True, size=10, color="CC0000" if val < 0 else "1A3C5E")
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[last].height = 18
        last += 1

    for col, w in zip(range(1, 6), [14, 40, 14, 40, 18]):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# HTML FRONTEND
# ─────────────────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Mauer · PDF</title>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
  :root {
    --bg:       #0d0f14;
    --surface:  #161920;
    --card:     #1c2030;
    --border:   #2a2e3d;
    --accent1:  #3b82f6;
    --accent2:  #10b981;
    --text:     #e2e8f0;
    --muted:    #64748b;
    --danger:   #ef4444;
    --radius:   12px;
  }

  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    font-family: 'DM Sans', sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    padding: 48px 24px 80px;
  }

  /* HEADER */
  .logo {
    font-family: 'Bebas Neue', sans-serif;
    font-size: 48px;
    letter-spacing: 4px;
    color: var(--text);
    margin-bottom: 4px;
  }
  .logo span { color: var(--accent1); }
  .tagline {
    font-size: 13px;
    color: var(--muted);
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 56px;
  }

  /* GRID */
  .grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    width: 100%;
    max-width: 860px;
  }

  /* CARDS */
  .card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 28px;
    display: flex;
    flex-direction: column;
    gap: 20px;
    transition: border-color .2s;
  }
  .card:hover { border-color: var(--accent1); }

  .card-header {
    display: flex;
    align-items: center;
    gap: 12px;
  }
  .card-icon {
    width: 40px; height: 40px;
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px;
    flex-shrink: 0;
  }
  .card-icon.blue  { background: rgba(59,130,246,.15); }
  .card-icon.green { background: rgba(16,185,129,.15); }

  .card-title {
    font-size: 16px;
    font-weight: 600;
    color: var(--text);
  }
  .card-sub {
    font-size: 12px;
    color: var(--muted);
    margin-top: 2px;
  }

  /* DROPZONE */
  .dropzone {
    border: 1.5px dashed var(--border);
    border-radius: 8px;
    padding: 28px 16px;
    text-align: center;
    cursor: pointer;
    transition: border-color .2s, background .2s;
    position: relative;
  }
  .dropzone:hover,
  .dropzone.dragover { border-color: var(--accent1); background: rgba(59,130,246,.05); }
  .dropzone.has-file { border-color: var(--accent2); background: rgba(16,185,129,.05); }
  .dropzone input { position: absolute; inset: 0; opacity: 0; cursor: pointer; }
  .dz-icon { font-size: 28px; margin-bottom: 8px; }
  .dz-text { font-size: 13px; color: var(--muted); }
  .dz-text strong { color: var(--text); }
  .dz-filename {
    font-size: 12px; color: var(--accent2);
    margin-top: 8px; font-weight: 500;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }

  /* BUTTON */
  .btn {
    width: 100%;
    padding: 13px;
    border: none;
    border-radius: 8px;
    font-family: 'DM Sans', sans-serif;
    font-size: 14px;
    font-weight: 600;
    letter-spacing: .5px;
    cursor: pointer;
    transition: opacity .2s, transform .1s;
    display: flex; align-items: center; justify-content: center; gap: 8px;
  }
  .btn:active { transform: scale(.98); }
  .btn:disabled { opacity: .4; cursor: not-allowed; }
  .btn-blue  { background: var(--accent1); color: #fff; }
  .btn-green { background: var(--accent2); color: #fff; }
  .password-input {
    width: 100%;
    padding: 12px 14px;
    border-radius: 10px;
    border: 1px solid var(--border);
    background: #12151f;
    color: var(--text);
    font-size: 14px;
    margin-top: 10px;
  }
  .password-input::placeholder { color: var(--muted); }

  /* STATUS */
  .status {
    width: 100%; max-width: 860px;
    margin-top: 20px;
    padding: 14px 18px;
    border-radius: 8px;
    font-size: 13px;
    display: none;
    align-items: center;
    gap: 10px;
  }
  .status.show { display: flex; }
  .status.loading { background: rgba(59,130,246,.1); border: 1px solid rgba(59,130,246,.3); color: #93c5fd; }
  .status.success { background: rgba(16,185,129,.1); border: 1px solid rgba(16,185,129,.3); color: #6ee7b7; }
  .status.error   { background: rgba(239,68,68,.1);  border: 1px solid rgba(239,68,68,.3);  color: #fca5a5; }

  /* SPINNER */
  @keyframes spin { to { transform: rotate(360deg); } }
  .spinner {
    width: 16px; height: 16px;
    border: 2px solid currentColor;
    border-top-color: transparent;
    border-radius: 50%;
    animation: spin .7s linear infinite;
    flex-shrink: 0;
  }

  /* LEGEND */
  .legend {
    width: 100%; max-width: 860px;
    margin-top: 32px;
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 10px;
  }
  .leg-item {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 12px 14px;
    font-size: 12px;
    color: var(--muted);
    display: flex; gap: 8px; align-items: flex-start;
  }
  .leg-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    margin-top: 3px; flex-shrink: 0;
  }

  @media (max-width: 640px) {
    .grid { grid-template-columns: 1fr; }
    .legend { grid-template-columns: 1fr; }
    .logo { font-size: 36px; }
  }
</style>
</head>
<body>

<div class="logo">MAUER <span>·</span> PDF</div>
<div class="tagline">Processador de Relatórios</div>

<div class="grid">

  <!-- SÓLIDES -->
  <div class="card">
    <div class="card-header">
      <div class="card-icon blue">⏱</div>
      <div>
        <div class="card-title">Folha de Ponto</div>
        <div class="card-sub">Sólides — RFP</div>
      </div>
    </div>

    <div class="dropzone" id="dz-solides">
      <input type="file" accept=".pdf" id="file-solides">
      <div class="dz-icon">📄</div>
      <div class="dz-text"><strong>Clique</strong> ou arraste o PDF aqui</div>
      <div class="dz-filename" id="fn-solides"></div>
    </div>
    <input class="password-input" type="password" id="pwd-solides" placeholder="Senha do PDF (opcional)">

    <button class="btn btn-blue" id="btn-solides" onclick="process('solides')" disabled>
      <span>Gerar Relatório Excel</span>
    </button>
  </div>

  <!-- BRADESCO -->
  <div class="card">
    <div class="card-header">
      <div class="card-icon green">🦷</div>
      <div>
        <div class="card-title">Fatura Dental</div>
        <div class="card-sub">Bradesco Saúde — Fatura Técnica</div>
      </div>
    </div>

    <div class="dropzone" id="dz-bradesco">
      <input type="file" accept=".pdf" id="file-bradesco">
      <div class="dz-icon">📄</div>
      <div class="dz-text"><strong>Clique</strong> ou arraste o PDF aqui</div>
      <div class="dz-filename" id="fn-bradesco"></div>
    </div>
    <input class="password-input" type="password" id="pwd-bradesco" placeholder="Senha do PDF (opcional)">

    <button class="btn btn-green" id="btn-bradesco" onclick="process('bradesco')" disabled>
      <span>Gerar Relatório Excel</span>
    </button>
  </div>

</div>

<div class="status" id="status">
  <div class="spinner" id="status-spinner" style="display:none"></div>
  <span id="status-text"></span>
</div>

<div class="legend">
  <div class="leg-item">
    <div class="leg-dot" style="background:#3b82f6"></div>
    <span><strong style="color:#e2e8f0">Horas Extras (> 1:12)</strong> — saldo mensal acima do limiar entra como hora extra</span>
  </div>
  <div class="leg-item">
    <div class="leg-dot" style="background:#f59e0b"></div>
    <span><strong style="color:#e2e8f0">Banco de Horas (≤ 1:12)</strong> — saldo positivo dentro do limiar vai pro banco</span>
  </div>
  <div class="leg-item">
    <div class="leg-dot" style="background:#ef4444"></div>
    <span><strong style="color:#e2e8f0">Saldo Negativo</strong> — colaborador com menos horas do que o previsto no período</span>
  </div>
</div>

<script>
  // Dropzone setup
  function setupDropzone(dzId, inputId, fnId, btnId) {
    const dz    = document.getElementById(dzId);
    const input = document.getElementById(inputId);
    const fn    = document.getElementById(fnId);
    const btn   = document.getElementById(btnId);

    input.addEventListener('change', () => {
      if (input.files[0]) {
        fn.textContent = input.files[0].name;
        dz.classList.add('has-file');
        btn.disabled = false;
      }
    });

    dz.addEventListener('dragover', e => { e.preventDefault(); dz.classList.add('dragover'); });
    dz.addEventListener('dragleave', () => dz.classList.remove('dragover'));
    dz.addEventListener('drop', e => {
      e.preventDefault();
      dz.classList.remove('dragover');
      const file = e.dataTransfer.files[0];
      if (file && file.name.endsWith('.pdf')) {
        const dt = new DataTransfer();
        dt.items.add(file);
        input.files = dt.files;
        fn.textContent = file.name;
        dz.classList.add('has-file');
        btn.disabled = false;
      }
    });
  }

  setupDropzone('dz-solides',  'file-solides',  'fn-solides',  'btn-solides');
  setupDropzone('dz-bradesco', 'file-bradesco', 'fn-bradesco', 'btn-bradesco');

  function setStatus(type, msg) {
    const el  = document.getElementById('status');
    const txt = document.getElementById('status-text');
    const sp  = document.getElementById('status-spinner');
    el.className = `status show ${type}`;
    txt.textContent = msg;
    sp.style.display = type === 'loading' ? 'block' : 'none';
  }

  async function process(type) {
    const inputId = type === 'solides' ? 'file-solides' : 'file-bradesco';
    const btnId   = type === 'solides' ? 'btn-solides'  : 'btn-bradesco';
    const file    = document.getElementById(inputId).files[0];
    if (!file) return;

    const btn = document.getElementById(btnId);
    btn.disabled = true;
    setStatus('loading', `Processando ${file.name}…`);

    const pwdId = type === 'solides' ? 'pwd-solides' : 'pwd-bradesco';
    const password = document.getElementById(pwdId).value || '';
    const form = new FormData();
    form.append('file', file);
    form.append('type', type);
    form.append('password', password);

    try {
      const res = await fetch('/process', { method: 'POST', body: form });
      if (!res.ok) {
        const err = await res.json();
        throw new Error(err.error || 'Erro desconhecido');
      }
      const blob = await res.blob();
      const cd = res.headers.get('Content-Disposition') || '';
      const match = cd.match(/filename="?([^"]+)"?/);
      const fname = match ? match[1] : 'relatorio.xlsx';
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url; a.download = fname; a.click();
      URL.revokeObjectURL(url);
      setStatus('success', `✓ ${fname} gerado e baixado com sucesso!`);
    } catch(e) {
      setStatus('error', `Erro: ${e.message}`);
    } finally {
      btn.disabled = false;
    }
  }
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/process", methods=["POST"])
def process():
    file = request.files.get("file")
    kind = request.form.get("type", "")
    password = request.form.get("password", "") or None

    if not file or not file.filename.endswith(".pdf"):
        return jsonify({"error": "Envie um arquivo PDF válido."}), 400

    pdf_bytes = file.read()

    try:
        if kind == "solides":
            xlsx_bytes = parse_solides(pdf_bytes, password)
            filename = Path(file.filename).stem + "_relatorio.xlsx"
        elif kind == "bradesco":
            xlsx_bytes = parse_bradesco(pdf_bytes, password)
            filename = Path(file.filename).stem + "_fatura.xlsx"
        else:
            return jsonify({"error": "Tipo inválido."}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("=" * 50)
    print("  Mauer · PDF")
    print(f"  http://localhost:{port}")
    print("=" * 50)
    app.run(debug=False, host="0.0.0.0", port=port)
